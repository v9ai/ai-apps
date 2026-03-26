"""
Scrapus Export Scheduler: Automated & Incremental Lead Export
==============================================================

Cron-like scheduling for daily/weekly exports with:
  - Incremental exports (only new/updated leads since last run)
  - Export history tracking in SQLite
  - File naming with timestamps and run IDs
  - Post-export validation (row counts, required fields, format checks)
  - Email-ready text summaries of new leads

Author: Scrapus Team
Target: Apple M1 16GB, zero cloud dependency
"""

import hashlib
import json
import logging
import os
import sqlite3
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Dict, List, Literal, Optional, Sequence, Tuple

from export_engine import (
    ExportFilter,
    ExportFormat,
    ExportLead,
    PaginationConfig,
    export_leads,
)
from crm_adapters import CRMTarget, export_for_crm

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

ScheduleFrequency = Literal["hourly", "daily", "weekly", "monthly"]


@dataclass
class ScheduleConfig:
    """Configuration for a scheduled export job."""
    job_id: str
    frequency: ScheduleFrequency = "daily"
    fmt: ExportFormat = "csv"
    crm_target: Optional[CRMTarget] = None  # if set, use CRM adapter instead
    output_dir: str = "./exports"
    file_prefix: str = "scrapus"
    filt: Optional[ExportFilter] = None
    incremental: bool = True  # only export new/updated since last run
    deduplicate_crm: bool = True
    enabled: bool = True
    # Email summary
    generate_summary: bool = True
    summary_dest: Optional[str] = None  # file path for summary text


@dataclass
class ExportRunRecord:
    """Record of a completed export run, stored in SQLite."""
    run_id: str
    job_id: str
    started_at: str
    completed_at: str
    status: str  # "success", "failed", "validated", "validation_failed"
    format: str
    crm_target: Optional[str]
    output_files: List[str]
    row_count: int
    lead_count: int
    incremental: bool
    watermark: Optional[str]  # ISO timestamp used as the lower bound
    validation_errors: List[str] = field(default_factory=list)
    summary_text: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        d["output_files"] = json.dumps(d["output_files"])
        d["validation_errors"] = json.dumps(d["validation_errors"])
        return d


# ---------------------------------------------------------------------------
# SQLite History Store
# ---------------------------------------------------------------------------

_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS export_runs (
    run_id              TEXT PRIMARY KEY,
    job_id              TEXT NOT NULL,
    started_at          TEXT NOT NULL,
    completed_at        TEXT NOT NULL,
    status              TEXT NOT NULL,
    format              TEXT NOT NULL,
    crm_target          TEXT,
    output_files        TEXT NOT NULL,  -- JSON array
    row_count           INTEGER NOT NULL,
    lead_count          INTEGER NOT NULL,
    incremental         INTEGER NOT NULL,
    watermark           TEXT,
    validation_errors   TEXT NOT NULL DEFAULT '[]',  -- JSON array
    summary_text        TEXT
);

CREATE INDEX IF NOT EXISTS idx_export_runs_job ON export_runs(job_id, completed_at);
CREATE INDEX IF NOT EXISTS idx_export_runs_status ON export_runs(status);

CREATE TABLE IF NOT EXISTS export_watermarks (
    job_id              TEXT PRIMARY KEY,
    last_watermark      TEXT NOT NULL,  -- ISO timestamp of last successful export
    updated_at          TEXT NOT NULL
);
"""


class ExportHistoryStore:
    """SQLite-backed export run history and watermark tracking."""

    def __init__(self, db_path: str = "./exports/export_history.db"):
        self.db_path = db_path
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        self._conn: Optional[sqlite3.Connection] = None
        self._init_schema()

    def _get_conn(self) -> sqlite3.Connection:
        if self._conn is None:
            self._conn = sqlite3.connect(self.db_path)
            self._conn.row_factory = sqlite3.Row
            self._conn.execute("PRAGMA journal_mode=WAL")
            self._conn.execute("PRAGMA synchronous=NORMAL")
        return self._conn

    def _init_schema(self) -> None:
        conn = self._get_conn()
        conn.executescript(_SCHEMA_SQL)
        conn.commit()

    def record_run(self, record: ExportRunRecord) -> None:
        """Insert a completed export run record."""
        conn = self._get_conn()
        d = record.to_dict()
        conn.execute(
            """
            INSERT OR REPLACE INTO export_runs
                (run_id, job_id, started_at, completed_at, status, format,
                 crm_target, output_files, row_count, lead_count,
                 incremental, watermark, validation_errors, summary_text)
            VALUES
                (:run_id, :job_id, :started_at, :completed_at, :status, :format,
                 :crm_target, :output_files, :row_count, :lead_count,
                 :incremental, :watermark, :validation_errors, :summary_text)
            """,
            d,
        )
        conn.commit()

    def update_watermark(self, job_id: str, watermark: str) -> None:
        """Set the incremental export watermark for a job."""
        conn = self._get_conn()
        conn.execute(
            """
            INSERT OR REPLACE INTO export_watermarks (job_id, last_watermark, updated_at)
            VALUES (?, ?, ?)
            """,
            (job_id, watermark, datetime.now().isoformat()),
        )
        conn.commit()

    def get_watermark(self, job_id: str) -> Optional[str]:
        """Get the last successful export watermark for a job."""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT last_watermark FROM export_watermarks WHERE job_id = ?",
            (job_id,),
        ).fetchone()
        return row["last_watermark"] if row else None

    def get_recent_runs(self, job_id: Optional[str] = None,
                        limit: int = 20) -> List[Dict[str, Any]]:
        """Retrieve recent export run records."""
        conn = self._get_conn()
        if job_id:
            rows = conn.execute(
                "SELECT * FROM export_runs WHERE job_id = ? ORDER BY completed_at DESC LIMIT ?",
                (job_id, limit),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM export_runs ORDER BY completed_at DESC LIMIT ?",
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]

    def get_run(self, run_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve a specific run by ID."""
        conn = self._get_conn()
        row = conn.execute(
            "SELECT * FROM export_runs WHERE run_id = ?", (run_id,)
        ).fetchone()
        return dict(row) if row else None

    def close(self) -> None:
        if self._conn:
            self._conn.close()
            self._conn = None


# ---------------------------------------------------------------------------
# Run ID Generation
# ---------------------------------------------------------------------------

def _generate_run_id(job_id: str) -> str:
    """Deterministic-ish run ID: job + timestamp + short hash."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    h = hashlib.sha256(f"{job_id}:{ts}:{os.getpid()}".encode()).hexdigest()[:8]
    return f"{job_id}_{ts}_{h}"


# ---------------------------------------------------------------------------
# Export Validation
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    """Result of post-export validation."""
    passed: bool
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    row_count: int = 0


def validate_export(output_path: Path,
                    fmt: ExportFormat,
                    expected_count: Optional[int] = None,
                    required_fields: Optional[List[str]] = None) -> ValidationResult:
    """
    Validate an exported file for correctness.

    Checks:
      1. File exists and is non-empty
      2. Row count matches expected (if provided)
      3. Required columns/fields are present
      4. Format-specific checks (valid CSV, valid JSON, etc.)
    """
    errors: List[str] = []
    warnings: List[str] = []
    row_count = 0

    if not output_path.exists():
        return ValidationResult(passed=False, errors=[f"File not found: {output_path}"])

    size = output_path.stat().st_size
    if size == 0:
        return ValidationResult(passed=False, errors=[f"File is empty: {output_path}"])

    if fmt == "csv":
        row_count, errs, warns = _validate_csv(output_path, required_fields)
        errors.extend(errs)
        warnings.extend(warns)

    elif fmt == "json":
        row_count, errs, warns = _validate_json(output_path, required_fields)
        errors.extend(errs)
        warnings.extend(warns)

    elif fmt == "jsonl":
        row_count, errs, warns = _validate_jsonl(output_path, required_fields)
        errors.extend(errs)
        warnings.extend(warns)

    elif fmt == "xlsx":
        row_count, errs, warns = _validate_xlsx(output_path)
        errors.extend(errs)
        warnings.extend(warns)

    elif fmt in ("html", "markdown"):
        # Lightweight check: file is non-empty and readable
        try:
            content = output_path.read_text(encoding="utf-8")
            if fmt == "html" and "<table" not in content.lower():
                warnings.append("HTML file does not contain a <table> element")
            row_count = content.count("\n")
        except Exception as e:
            errors.append(f"Cannot read file: {e}")

    if expected_count is not None and row_count != expected_count:
        errors.append(
            f"Row count mismatch: expected {expected_count}, got {row_count}"
        )

    passed = len(errors) == 0
    return ValidationResult(passed=passed, errors=errors, warnings=warnings, row_count=row_count)


def _validate_csv(path: Path,
                  required_fields: Optional[List[str]]) -> Tuple[int, List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    row_count = 0
    try:
        import csv as csv_mod
        with open(path, "r", encoding="utf-8") as fh:
            reader = csv_mod.DictReader(fh)
            headers = reader.fieldnames or []
            if required_fields:
                missing = set(required_fields) - set(headers)
                if missing:
                    errors.append(f"Missing required CSV columns: {missing}")
            for row in reader:
                row_count += 1
    except Exception as e:
        errors.append(f"CSV parse error: {e}")
    return row_count, errors, warnings


def _validate_json(path: Path,
                   required_fields: Optional[List[str]]) -> Tuple[int, List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    row_count = 0
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if isinstance(data, dict) and "leads" in data:
            leads = data["leads"]
            row_count = len(leads)
            if required_fields and leads:
                missing = set(required_fields) - set(leads[0].keys())
                if missing:
                    errors.append(f"Missing required JSON fields: {missing}")
        else:
            warnings.append("JSON structure does not contain 'leads' key")
    except json.JSONDecodeError as e:
        errors.append(f"Invalid JSON: {e}")
    except Exception as e:
        errors.append(f"JSON read error: {e}")
    return row_count, errors, warnings


def _validate_jsonl(path: Path,
                    required_fields: Optional[List[str]]) -> Tuple[int, List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    row_count = 0
    try:
        with open(path, "r", encoding="utf-8") as fh:
            for lineno, line in enumerate(fh, start=1):
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                    row_count += 1
                    if required_fields and lineno == 1:
                        missing = set(required_fields) - set(obj.keys())
                        if missing:
                            errors.append(f"Missing required JSONL fields (line 1): {missing}")
                except json.JSONDecodeError as e:
                    errors.append(f"Invalid JSON on line {lineno}: {e}")
                    if len(errors) > 10:
                        errors.append("... (truncated, too many errors)")
                        break
    except Exception as e:
        errors.append(f"JSONL read error: {e}")
    return row_count, errors, warnings


def _validate_xlsx(path: Path) -> Tuple[int, List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    row_count = 0
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True)
        if "Leads" in wb.sheetnames:
            ws = wb["Leads"]
            row_count = ws.max_row - 1 if ws.max_row else 0  # minus header
        else:
            warnings.append("XLSX file has no 'Leads' sheet")
        wb.close()
    except ImportError:
        warnings.append("openpyxl not installed; skipping XLSX validation")
    except Exception as e:
        errors.append(f"XLSX read error: {e}")
    return row_count, errors, warnings


# ---------------------------------------------------------------------------
# Incremental Filtering
# ---------------------------------------------------------------------------

def filter_incremental(leads: Sequence[ExportLead],
                       watermark: Optional[str]) -> List[ExportLead]:
    """
    Return only leads created or updated after the watermark timestamp.
    If no watermark, return all leads (first run).
    """
    if watermark is None:
        return list(leads)

    result: List[ExportLead] = []
    for lead in leads:
        ts = lead.updated_at or lead.created_at
        if ts is None or ts > watermark:
            result.append(lead)
    return result


def compute_new_watermark(leads: Sequence[ExportLead]) -> str:
    """Compute the maximum updated_at / created_at across exported leads."""
    max_ts = ""
    for lead in leads:
        ts = lead.updated_at or lead.created_at or ""
        if ts > max_ts:
            max_ts = ts
    return max_ts or datetime.now().isoformat()


# ---------------------------------------------------------------------------
# Email-Ready Summary
# ---------------------------------------------------------------------------

def generate_lead_summary(leads: Sequence[ExportLead],
                          run_id: str,
                          incremental: bool = False) -> str:
    """
    Generate a plain-text summary of exported leads suitable for email
    or Slack notifications.
    """
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    qualified = [l for l in leads if l.is_qualified]
    scores = [l.lead_score for l in leads]
    avg_score = sum(scores) / len(scores) if scores else 0.0

    lines: List[str] = []
    lines.append(f"SCRAPUS LEAD EXPORT SUMMARY")
    lines.append(f"===========================")
    lines.append(f"Run ID:       {run_id}")
    lines.append(f"Generated:    {now}")
    lines.append(f"Type:         {'Incremental' if incremental else 'Full'}")
    lines.append(f"Total leads:  {len(leads)}")
    lines.append(f"Qualified:    {len(qualified)}")
    lines.append(f"Avg score:    {avg_score:.4f}")
    lines.append("")

    if qualified:
        lines.append("TOP QUALIFIED LEADS:")
        lines.append("-" * 40)
        top = sorted(qualified, key=lambda l: l.lead_score, reverse=True)[:10]
        for i, lead in enumerate(top, start=1):
            lines.append(
                f"  {i}. {lead.company_name} "
                f"(score={lead.lead_score:.3f}, conf={lead.lead_confidence:.3f})"
            )
            if lead.industry:
                lines.append(f"     Industry: {lead.industry}")
            if lead.contacts:
                c = lead.contacts[0]
                lines.append(f"     Contact:  {c.name} ({c.email or 'no email'})")
            if lead.report:
                lines.append(f"     Summary:  {lead.report.summary[:120]}...")
            lines.append("")

    # Industry breakdown
    industry_counts: Dict[str, int] = {}
    for l in leads:
        key = l.industry or "Unknown"
        industry_counts[key] = industry_counts.get(key, 0) + 1
    if industry_counts:
        lines.append("INDUSTRY BREAKDOWN:")
        lines.append("-" * 40)
        for ind, cnt in sorted(industry_counts.items(), key=lambda x: -x[1])[:10]:
            lines.append(f"  {ind}: {cnt} leads")
        lines.append("")

    lines.append(f"--- End of summary ({run_id}) ---")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Schedule Checker
# ---------------------------------------------------------------------------

def is_due(config: ScheduleConfig, history: ExportHistoryStore) -> bool:
    """
    Check if a scheduled export job is due to run based on its frequency
    and the timestamp of its last successful run.
    """
    if not config.enabled:
        return False

    recent = history.get_recent_runs(job_id=config.job_id, limit=1)
    if not recent:
        return True  # never run before

    last_run_ts = recent[0].get("completed_at", "")
    if not last_run_ts:
        return True

    try:
        last_run = datetime.fromisoformat(last_run_ts)
    except ValueError:
        return True

    now = datetime.now()
    intervals = {
        "hourly": timedelta(hours=1),
        "daily": timedelta(days=1),
        "weekly": timedelta(weeks=1),
        "monthly": timedelta(days=30),
    }
    interval = intervals.get(config.frequency, timedelta(days=1))
    return now - last_run >= interval


# ---------------------------------------------------------------------------
# Executor
# ---------------------------------------------------------------------------

def execute_export_job(config: ScheduleConfig,
                       leads: Sequence[ExportLead],
                       history: ExportHistoryStore,
                       force: bool = False) -> Optional[ExportRunRecord]:
    """
    Execute a single scheduled export job.

    Steps:
      1. Check if due (or force=True)
      2. Apply incremental watermark filtering
      3. Run export (file or CRM)
      4. Validate output
      5. Record in history
      6. Update watermark
      7. Generate summary

    Returns:
        ExportRunRecord if export was executed, None if skipped.
    """
    if not force and not is_due(config, history):
        logger.debug("Job %s is not due yet, skipping", config.job_id)
        return None

    run_id = _generate_run_id(config.job_id)
    started_at = datetime.now().isoformat()
    logger.info("Starting export job: %s (run_id=%s)", config.job_id, run_id)

    # Incremental filtering
    watermark: Optional[str] = None
    export_leads_list: List[ExportLead]
    if config.incremental:
        watermark = history.get_watermark(config.job_id)
        export_leads_list = filter_incremental(leads, watermark)
        logger.info(
            "Incremental filter: watermark=%s, %d/%d leads selected",
            watermark, len(export_leads_list), len(leads),
        )
    else:
        export_leads_list = list(leads)

    if not export_leads_list:
        logger.info("Job %s: no new leads to export", config.job_id)
        record = ExportRunRecord(
            run_id=run_id,
            job_id=config.job_id,
            started_at=started_at,
            completed_at=datetime.now().isoformat(),
            status="success",
            format=config.fmt,
            crm_target=config.crm_target,
            output_files=[],
            row_count=0,
            lead_count=0,
            incremental=config.incremental,
            watermark=watermark,
            summary_text="No new leads to export.",
        )
        history.record_run(record)
        return record

    # Build output path
    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_files: List[str] = []
    total_rows = 0

    try:
        if config.crm_target:
            # CRM adapter export
            crm_results = export_for_crm(
                export_leads_list,
                crm=config.crm_target,
                output_dir=output_dir,
                filt=config.filt,
                deduplicate=config.deduplicate_crm,
                prefix=config.file_prefix,
            )
            for fname, count in crm_results.items():
                output_files.append(str(output_dir / fname))
                total_rows += count
        else:
            # Standard format export
            ext_map = {
                "csv": ".csv", "json": ".json", "jsonl": ".jsonl",
                "xlsx": ".xlsx", "html": ".html", "markdown": ".md",
            }
            ext = ext_map.get(config.fmt, ".csv")
            filename = f"{config.file_prefix}_{ts}{ext}"
            output_path = output_dir / filename
            total_rows = export_leads(
                export_leads_list,
                dest=output_path,
                fmt=config.fmt,
                filt=config.filt,
            )
            output_files.append(str(output_path))

        status = "success"

    except Exception as e:
        logger.error("Export job %s failed: %s", config.job_id, e, exc_info=True)
        record = ExportRunRecord(
            run_id=run_id,
            job_id=config.job_id,
            started_at=started_at,
            completed_at=datetime.now().isoformat(),
            status="failed",
            format=config.fmt,
            crm_target=config.crm_target,
            output_files=output_files,
            row_count=0,
            lead_count=len(export_leads_list),
            incremental=config.incremental,
            watermark=watermark,
            validation_errors=[str(e)],
        )
        history.record_run(record)
        return record

    # Validation
    validation_errors: List[str] = []
    if not config.crm_target:
        for fpath in output_files:
            vr = validate_export(
                Path(fpath),
                fmt=config.fmt,
                expected_count=total_rows,
                required_fields=["company_id", "company_name"] if config.fmt == "csv" else None,
            )
            if not vr.passed:
                validation_errors.extend(vr.errors)
                status = "validation_failed"
            if vr.warnings:
                for w in vr.warnings:
                    logger.warning("Validation warning (%s): %s", fpath, w)

    # Summary
    summary_text: Optional[str] = None
    if config.generate_summary:
        summary_text = generate_lead_summary(
            export_leads_list, run_id, incremental=config.incremental
        )
        if config.summary_dest:
            summary_path = Path(config.summary_dest)
            summary_path.parent.mkdir(parents=True, exist_ok=True)
            summary_path.write_text(summary_text, encoding="utf-8")
            logger.info("Summary written to %s", summary_path)

    # New watermark
    new_watermark = compute_new_watermark(export_leads_list)

    # Record
    record = ExportRunRecord(
        run_id=run_id,
        job_id=config.job_id,
        started_at=started_at,
        completed_at=datetime.now().isoformat(),
        status=status,
        format=config.fmt,
        crm_target=config.crm_target,
        output_files=output_files,
        row_count=total_rows,
        lead_count=len(export_leads_list),
        incremental=config.incremental,
        watermark=new_watermark,
        validation_errors=validation_errors,
        summary_text=summary_text,
    )
    history.record_run(record)

    # Update watermark only on success
    if status in ("success", "validated"):
        history.update_watermark(config.job_id, new_watermark)

    logger.info(
        "Export job complete: job_id=%s, run_id=%s, status=%s, rows=%d, files=%d",
        config.job_id, run_id, status, total_rows, len(output_files),
    )
    return record


# ---------------------------------------------------------------------------
# Batch Scheduler (run all due jobs)
# ---------------------------------------------------------------------------

def run_scheduled_exports(configs: Sequence[ScheduleConfig],
                          leads: Sequence[ExportLead],
                          history: ExportHistoryStore,
                          force_all: bool = False) -> List[ExportRunRecord]:
    """
    Check all scheduled jobs and execute those that are due.

    Args:
        configs: List of job configurations.
        leads: Full lead set (incremental filtering applied per-job).
        history: SQLite history store.
        force_all: If True, run all jobs regardless of schedule.

    Returns:
        List of ExportRunRecord for all executed jobs.
    """
    results: List[ExportRunRecord] = []
    for config in configs:
        if not config.enabled:
            logger.debug("Job %s is disabled, skipping", config.job_id)
            continue
        record = execute_export_job(config, leads, history, force=force_all)
        if record is not None:
            results.append(record)
    logger.info(
        "Scheduled export batch: %d/%d jobs executed",
        len(results), len(configs),
    )
    return results
