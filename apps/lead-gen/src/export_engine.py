"""
Scrapus Export Engine: Multi-Format Lead Data Export
=====================================================

Multi-format export engine for Scrapus pipeline output.
Supports CSV, JSON, JSONL, Excel/XLSX, HTML, and Markdown exports
with filtering, pagination, and template customization.

Author: Scrapus Team
Target: Apple M1 16GB, zero cloud dependency
"""

import csv
import io
import json
import logging
import math
import os
import sqlite3
from dataclasses import dataclass, field, asdict
from datetime import datetime, date
from pathlib import Path
from typing import (
    Any, Callable, Dict, Iterator, List, Literal, Optional, Sequence, Tuple, Union,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data Structures
# ---------------------------------------------------------------------------

@dataclass
class ExportContact:
    """Contact record extracted from pipeline entities."""
    contact_id: int
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    company_id: Optional[int] = None
    company_name: Optional[str] = None
    linkedin_url: Optional[str] = None
    source_url: Optional[str] = None
    created_at: Optional[str] = None


@dataclass
class ExportEntity:
    """Resolved entity from Module 3."""
    entity_id: int
    entity_type: str  # ORG, PERSON, LOCATION, PRODUCT, FUNDING_ROUND, TECHNOLOGY
    name: str
    normalized_name: Optional[str] = None
    confidence: float = 0.0
    source_url: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


@dataclass
class ExportReport:
    """Generated report from Module 5 (structured_output.py LeadReport)."""
    report_id: int
    company_id: int
    summary: str = ""
    key_strengths: List[str] = field(default_factory=list)
    growth_indicators: List[str] = field(default_factory=list)
    risk_factors: List[str] = field(default_factory=list)
    recommended_approach: str = ""
    confidence: float = 0.0
    sources: List[Dict[str, Any]] = field(default_factory=list)
    generation_timestamp: Optional[str] = None


@dataclass
class SHAPExplanation:
    """SHAP feature-importance explanation for a lead score."""
    feature_name: str
    shap_value: float
    feature_value: float


@dataclass
class ConformalInterval:
    """Conformal prediction interval from conformal_pipeline.py."""
    lower_bound: float
    upper_bound: float
    interval_width: float
    coverage_level: float = 0.95
    method: str = "plus"


@dataclass
class ExportLead:
    """Fully-hydrated lead record ready for export."""
    company_id: int
    company_name: str
    domain: Optional[str] = None
    industry: Optional[str] = None
    size_tier: Optional[str] = None
    location: Optional[str] = None
    employee_count: Optional[int] = None
    funding_amount_usd: Optional[float] = None
    lead_score: float = 0.0
    lead_confidence: float = 0.0
    is_qualified: bool = False
    qualification_reason: Optional[str] = None
    conformal: Optional[ConformalInterval] = None
    shap_explanations: List[SHAPExplanation] = field(default_factory=list)
    contacts: List[ExportContact] = field(default_factory=list)
    entities: List[ExportEntity] = field(default_factory=list)
    report: Optional[ExportReport] = None
    source_urls: List[str] = field(default_factory=list)
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


# ---------------------------------------------------------------------------
# Filtering & Pagination
# ---------------------------------------------------------------------------

@dataclass
class ExportFilter:
    """Declarative filter for lead exports."""
    qualified_only: bool = False
    min_score: Optional[float] = None
    max_score: Optional[float] = None
    min_confidence: Optional[float] = None
    industries: Optional[List[str]] = None
    size_tiers: Optional[List[str]] = None
    entity_types: Optional[List[str]] = None
    date_from: Optional[str] = None   # ISO date YYYY-MM-DD
    date_to: Optional[str] = None
    has_report: Optional[bool] = None
    has_contacts: Optional[bool] = None
    search_term: Optional[str] = None  # fuzzy match on company_name / domain


@dataclass
class PaginationConfig:
    """Pagination settings for large exports."""
    page_size: int = 500
    page_number: int = 1  # 1-indexed

    @property
    def offset(self) -> int:
        return (self.page_number - 1) * self.page_size


def apply_filters(leads: Sequence[ExportLead],
                  filt: Optional[ExportFilter] = None) -> List[ExportLead]:
    """Apply declarative filters to a lead sequence, returning a new list."""
    if filt is None:
        return list(leads)

    result: List[ExportLead] = []
    for lead in leads:
        if filt.qualified_only and not lead.is_qualified:
            continue
        if filt.min_score is not None and lead.lead_score < filt.min_score:
            continue
        if filt.max_score is not None and lead.lead_score > filt.max_score:
            continue
        if filt.min_confidence is not None and lead.lead_confidence < filt.min_confidence:
            continue
        if filt.industries and (lead.industry or "").lower() not in [
            i.lower() for i in filt.industries
        ]:
            continue
        if filt.size_tiers and (lead.size_tier or "").lower() not in [
            s.lower() for s in filt.size_tiers
        ]:
            continue
        if filt.entity_types:
            types_present = {e.entity_type for e in lead.entities}
            if not types_present.intersection(set(filt.entity_types)):
                continue
        if filt.date_from:
            if lead.created_at and lead.created_at < filt.date_from:
                continue
        if filt.date_to:
            if lead.created_at and lead.created_at > filt.date_to:
                continue
        if filt.has_report is True and lead.report is None:
            continue
        if filt.has_report is False and lead.report is not None:
            continue
        if filt.has_contacts is True and not lead.contacts:
            continue
        if filt.has_contacts is False and lead.contacts:
            continue
        if filt.search_term:
            term = filt.search_term.lower()
            haystack = f"{lead.company_name} {lead.domain or ''}".lower()
            if term not in haystack:
                continue
        result.append(lead)
    return result


def paginate(leads: Sequence[ExportLead],
             config: Optional[PaginationConfig] = None) -> List[ExportLead]:
    """Return a single page of leads."""
    if config is None:
        return list(leads)
    start = config.offset
    end = start + config.page_size
    return list(leads[start:end])


def total_pages(total: int, page_size: int) -> int:
    return max(1, math.ceil(total / page_size))


# ---------------------------------------------------------------------------
# Serialization Helpers
# ---------------------------------------------------------------------------

def _lead_flat_row(lead: ExportLead) -> Dict[str, Any]:
    """Flatten a lead into a single-level dict suitable for CSV / tabular."""
    row: Dict[str, Any] = {
        "company_id": lead.company_id,
        "company_name": lead.company_name,
        "domain": lead.domain or "",
        "industry": lead.industry or "",
        "size_tier": lead.size_tier or "",
        "location": lead.location or "",
        "employee_count": lead.employee_count or "",
        "funding_amount_usd": lead.funding_amount_usd or "",
        "lead_score": round(lead.lead_score, 4),
        "lead_confidence": round(lead.lead_confidence, 4),
        "is_qualified": lead.is_qualified,
        "qualification_reason": lead.qualification_reason or "",
    }
    # Conformal interval columns
    if lead.conformal:
        row["conformal_lower"] = round(lead.conformal.lower_bound, 4)
        row["conformal_upper"] = round(lead.conformal.upper_bound, 4)
        row["conformal_width"] = round(lead.conformal.interval_width, 4)
        row["conformal_coverage"] = lead.conformal.coverage_level
    else:
        row["conformal_lower"] = ""
        row["conformal_upper"] = ""
        row["conformal_width"] = ""
        row["conformal_coverage"] = ""
    # Top SHAP feature
    if lead.shap_explanations:
        top = max(lead.shap_explanations, key=lambda s: abs(s.shap_value))
        row["top_shap_feature"] = top.feature_name
        row["top_shap_value"] = round(top.shap_value, 4)
    else:
        row["top_shap_feature"] = ""
        row["top_shap_value"] = ""
    # First contact
    if lead.contacts:
        c = lead.contacts[0]
        row["primary_contact_name"] = c.name
        row["primary_contact_email"] = c.email or ""
        row["primary_contact_phone"] = c.phone or ""
        row["primary_contact_role"] = c.role or ""
    else:
        row["primary_contact_name"] = ""
        row["primary_contact_email"] = ""
        row["primary_contact_phone"] = ""
        row["primary_contact_role"] = ""
    row["contact_count"] = len(lead.contacts)
    row["entity_count"] = len(lead.entities)
    row["has_report"] = lead.report is not None
    row["source_urls"] = "; ".join(lead.source_urls) if lead.source_urls else ""
    row["created_at"] = lead.created_at or ""
    row["updated_at"] = lead.updated_at or ""
    return row


def _lead_full_dict(lead: ExportLead) -> Dict[str, Any]:
    """Deep-serialize a lead to a JSON-compatible dict with nested objects."""
    d: Dict[str, Any] = {
        "company_id": lead.company_id,
        "company_name": lead.company_name,
        "domain": lead.domain,
        "industry": lead.industry,
        "size_tier": lead.size_tier,
        "location": lead.location,
        "employee_count": lead.employee_count,
        "funding_amount_usd": lead.funding_amount_usd,
        "lead_score": lead.lead_score,
        "lead_confidence": lead.lead_confidence,
        "is_qualified": lead.is_qualified,
        "qualification_reason": lead.qualification_reason,
        "created_at": lead.created_at,
        "updated_at": lead.updated_at,
        "source_urls": lead.source_urls,
    }
    if lead.conformal:
        d["conformal_interval"] = asdict(lead.conformal)
    else:
        d["conformal_interval"] = None
    d["shap_explanations"] = [asdict(s) for s in lead.shap_explanations]
    d["contacts"] = [asdict(c) for c in lead.contacts]
    d["entities"] = [
        {**asdict(e), "metadata": e.metadata} for e in lead.entities
    ]
    if lead.report:
        d["report"] = asdict(lead.report)
    else:
        d["report"] = None
    return d


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------

def export_csv(leads: Sequence[ExportLead],
               dest: Union[str, Path, io.StringIO],
               filt: Optional[ExportFilter] = None,
               pagination: Optional[PaginationConfig] = None) -> int:
    """
    Export leads as CSV with flat columns for scores, contacts, conformal intervals.

    Returns:
        Number of rows written.
    """
    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)
    if not page:
        logger.warning("CSV export: zero leads after filtering")
        return 0

    rows = [_lead_flat_row(l) for l in page]
    fieldnames = list(rows[0].keys())

    close_after = False
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        fh = open(dest, "w", newline="", encoding="utf-8")
        close_after = True
    else:
        fh = dest

    try:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    finally:
        if close_after:
            fh.close()

    logger.info("CSV export: %d rows written", len(rows))
    return len(rows)


# ---------------------------------------------------------------------------
# JSON Export
# ---------------------------------------------------------------------------

def export_json(leads: Sequence[ExportLead],
                dest: Union[str, Path, io.StringIO],
                filt: Optional[ExportFilter] = None,
                pagination: Optional[PaginationConfig] = None,
                indent: int = 2) -> int:
    """
    Export leads as a single JSON document with nested entities, reports, SHAP.

    Returns:
        Number of leads written.
    """
    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)
    payload = {
        "export_timestamp": datetime.now().isoformat(),
        "total_leads": len(filtered),
        "page": pagination.page_number if pagination else 1,
        "page_size": pagination.page_size if pagination else len(filtered),
        "total_pages": total_pages(len(filtered), pagination.page_size) if pagination else 1,
        "leads": [_lead_full_dict(l) for l in page],
    }

    close_after = False
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        fh = open(dest, "w", encoding="utf-8")
        close_after = True
    else:
        fh = dest

    try:
        json.dump(payload, fh, indent=indent, default=str, ensure_ascii=False)
    finally:
        if close_after:
            fh.close()

    logger.info("JSON export: %d leads written", len(page))
    return len(page)


# ---------------------------------------------------------------------------
# JSON Lines (streaming) Export
# ---------------------------------------------------------------------------

def export_jsonl(leads: Sequence[ExportLead],
                 dest: Union[str, Path, io.StringIO],
                 filt: Optional[ExportFilter] = None) -> int:
    """
    Export leads as JSON Lines -- one JSON object per line.
    Suitable for streaming ingestion of large datasets without loading full file.

    Returns:
        Number of lines written.
    """
    filtered = apply_filters(leads, filt)
    if not filtered:
        logger.warning("JSONL export: zero leads after filtering")
        return 0

    close_after = False
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        fh = open(dest, "w", encoding="utf-8")
        close_after = True
    else:
        fh = dest

    count = 0
    try:
        for lead in filtered:
            line = json.dumps(_lead_full_dict(lead), default=str, ensure_ascii=False)
            fh.write(line + "\n")
            count += 1
    finally:
        if close_after:
            fh.close()

    logger.info("JSONL export: %d lines written", count)
    return count


# ---------------------------------------------------------------------------
# Excel / XLSX Export (via openpyxl)
# ---------------------------------------------------------------------------

def export_xlsx(leads: Sequence[ExportLead],
                dest: Union[str, Path],
                filt: Optional[ExportFilter] = None,
                pagination: Optional[PaginationConfig] = None) -> int:
    """
    Export leads as a formatted Excel workbook with four sheets:
      - Leads: main lead table
      - Contacts: all contacts across leads
      - Reports: generated reports per lead
      - Stats: summary statistics

    Returns:
        Number of lead rows written.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        )

    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)

    wb = Workbook()

    # -- Styles -----------------------------------------------------------------
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    qualified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    not_qualified_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    score_high_font = Font(name="Calibri", bold=True, color="006100")
    score_low_font = Font(name="Calibri", color="9C0006")

    def _style_header(ws, col_count: int) -> None:
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(col_count)}1"

    # -- Sheet 1: Leads ---------------------------------------------------------
    ws_leads = wb.active
    ws_leads.title = "Leads"
    lead_headers = [
        "Company ID", "Company Name", "Domain", "Industry", "Size Tier",
        "Location", "Employees", "Funding (USD)", "Lead Score",
        "Confidence", "Qualified", "Qualification Reason",
        "Conformal Lower", "Conformal Upper", "Conformal Width",
        "Top SHAP Feature", "Top SHAP Value",
        "Contact Count", "Entity Count", "Has Report",
        "Source URLs", "Created At", "Updated At",
    ]
    ws_leads.append(lead_headers)
    _style_header(ws_leads, len(lead_headers))

    for idx, lead in enumerate(page, start=2):
        top_shap_name = ""
        top_shap_val = ""
        if lead.shap_explanations:
            top = max(lead.shap_explanations, key=lambda s: abs(s.shap_value))
            top_shap_name = top.feature_name
            top_shap_val = round(top.shap_value, 4)
        conf_lo = round(lead.conformal.lower_bound, 4) if lead.conformal else ""
        conf_hi = round(lead.conformal.upper_bound, 4) if lead.conformal else ""
        conf_w = round(lead.conformal.interval_width, 4) if lead.conformal else ""

        row_data = [
            lead.company_id, lead.company_name, lead.domain or "",
            lead.industry or "", lead.size_tier or "", lead.location or "",
            lead.employee_count or "", lead.funding_amount_usd or "",
            round(lead.lead_score, 4), round(lead.lead_confidence, 4),
            "Yes" if lead.is_qualified else "No",
            lead.qualification_reason or "",
            conf_lo, conf_hi, conf_w,
            top_shap_name, top_shap_val,
            len(lead.contacts), len(lead.entities),
            "Yes" if lead.report else "No",
            "; ".join(lead.source_urls) if lead.source_urls else "",
            lead.created_at or "", lead.updated_at or "",
        ]
        ws_leads.append(row_data)

        # Conditional formatting
        qual_cell = ws_leads.cell(row=idx, column=11)
        if lead.is_qualified:
            qual_cell.fill = qualified_fill
        else:
            qual_cell.fill = not_qualified_fill

        score_cell = ws_leads.cell(row=idx, column=9)
        if lead.lead_score >= 0.7:
            score_cell.font = score_high_font
        elif lead.lead_score < 0.4:
            score_cell.font = score_low_font

    # Auto-width for lead columns
    for col_idx in range(1, len(lead_headers) + 1):
        max_len = len(str(lead_headers[col_idx - 1]))
        for row_idx in range(2, min(len(page) + 2, 52)):  # sample first 50
            val = ws_leads.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), 50))
        ws_leads.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # -- Sheet 2: Contacts ------------------------------------------------------
    ws_contacts = wb.create_sheet("Contacts")
    contact_headers = [
        "Contact ID", "Name", "Email", "Phone", "Role",
        "Company ID", "Company Name", "LinkedIn", "Source URL", "Created At",
    ]
    ws_contacts.append(contact_headers)
    _style_header(ws_contacts, len(contact_headers))
    for lead in page:
        for c in lead.contacts:
            ws_contacts.append([
                c.contact_id, c.name, c.email or "", c.phone or "",
                c.role or "", c.company_id or lead.company_id,
                c.company_name or lead.company_name,
                c.linkedin_url or "", c.source_url or "", c.created_at or "",
            ])

    # -- Sheet 3: Reports -------------------------------------------------------
    ws_reports = wb.create_sheet("Reports")
    report_headers = [
        "Company ID", "Company Name", "Summary",
        "Key Strengths", "Growth Indicators", "Risk Factors",
        "Recommended Approach", "Report Confidence", "Sources",
        "Generated At",
    ]
    ws_reports.append(report_headers)
    _style_header(ws_reports, len(report_headers))
    for lead in page:
        if lead.report:
            r = lead.report
            ws_reports.append([
                lead.company_id, lead.company_name, r.summary,
                " | ".join(r.key_strengths),
                " | ".join(r.growth_indicators),
                " | ".join(r.risk_factors),
                r.recommended_approach, round(r.confidence, 4),
                "; ".join(s.get("url", "") for s in r.sources),
                r.generation_timestamp or "",
            ])

    # -- Sheet 4: Stats ----------------------------------------------------------
    ws_stats = wb.create_sheet("Stats")
    stats_headers = ["Metric", "Value"]
    ws_stats.append(stats_headers)
    _style_header(ws_stats, len(stats_headers))

    total_filtered = len(filtered)
    qualified_count = sum(1 for l in filtered if l.is_qualified)
    scores = [l.lead_score for l in filtered]
    avg_score = sum(scores) / len(scores) if scores else 0.0
    max_score = max(scores) if scores else 0.0
    min_score = min(scores) if scores else 0.0
    reports_count = sum(1 for l in filtered if l.report is not None)
    contacts_count = sum(len(l.contacts) for l in filtered)
    entities_count = sum(len(l.entities) for l in filtered)

    stats_rows = [
        ("Export Timestamp", datetime.now().isoformat()),
        ("Total Leads (filtered)", total_filtered),
        ("Leads on This Page", len(page)),
        ("Qualified Leads", qualified_count),
        ("Qualification Rate", f"{(qualified_count / total_filtered * 100) if total_filtered else 0:.1f}%"),
        ("Avg Lead Score", f"{avg_score:.4f}"),
        ("Min Lead Score", f"{min_score:.4f}"),
        ("Max Lead Score", f"{max_score:.4f}"),
        ("Leads with Reports", reports_count),
        ("Total Contacts", contacts_count),
        ("Total Entities", entities_count),
    ]
    # Industry breakdown
    industry_counts: Dict[str, int] = {}
    for l in filtered:
        key = l.industry or "Unknown"
        industry_counts[key] = industry_counts.get(key, 0) + 1
    for ind, cnt in sorted(industry_counts.items(), key=lambda x: -x[1]):
        stats_rows.append((f"Industry: {ind}", cnt))

    for metric, value in stats_rows:
        ws_stats.append([metric, value])

    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 25

    # -- Save -------------------------------------------------------------------
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
    logger.info("XLSX export: %d leads written to %s", len(page), dest)
    return len(page)


# ---------------------------------------------------------------------------
# HTML Report Export
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Scrapus Lead Export</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         background: #f5f6fa; color: #2d3436; padding: 24px; }}
  h1 {{ font-size: 1.6rem; margin-bottom: 8px; color: #2f3640; }}
  .meta {{ color: #636e72; font-size: 0.85rem; margin-bottom: 20px; }}
  .stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 24px; }}
  .stat-card {{ background: #fff; border-radius: 8px; padding: 16px 20px;
                box-shadow: 0 1px 3px rgba(0,0,0,.08); min-width: 140px; }}
  .stat-card .label {{ font-size: 0.75rem; text-transform: uppercase;
                       letter-spacing: 0.5px; color: #636e72; }}
  .stat-card .value {{ font-size: 1.5rem; font-weight: 700; color: #2f5496; }}
  table {{ width: 100%; border-collapse: collapse; background: #fff;
           border-radius: 8px; overflow: hidden;
           box-shadow: 0 1px 3px rgba(0,0,0,.08); margin-bottom: 24px; }}
  th {{ background: #2f5496; color: #fff; padding: 10px 12px; font-size: 0.8rem;
       text-transform: uppercase; letter-spacing: 0.3px; cursor: pointer;
       user-select: none; white-space: nowrap; text-align: left; }}
  th:hover {{ background: #1e3a6e; }}
  td {{ padding: 8px 12px; font-size: 0.85rem; border-bottom: 1px solid #f0f0f0;
       max-width: 260px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  tr:hover {{ background: #f8f9ff; }}
  .qualified {{ color: #006100; font-weight: 600; }}
  .not-qualified {{ color: #9c0006; }}
  .score-high {{ font-weight: 700; color: #006100; }}
  .score-low {{ color: #9c0006; }}
  .details {{ display: none; background: #fafbfc; }}
  .details td {{ white-space: normal; max-width: none; padding: 16px; }}
  .details.open {{ display: table-row; }}
  .toggle {{ cursor: pointer; color: #2f5496; font-size: 0.8rem; text-decoration: underline; }}
  .report-section {{ margin-top: 8px; }}
  .report-section h4 {{ font-size: 0.85rem; color: #2f5496; margin-bottom: 4px; }}
  .report-section ul {{ padding-left: 18px; margin-bottom: 8px; }}
  .report-section li {{ font-size: 0.82rem; margin-bottom: 2px; }}
</style>
</head>
<body>
<h1>Scrapus Lead Export</h1>
<p class="meta">Generated {timestamp} &middot; {total_leads} leads</p>
<div class="stats">
  <div class="stat-card"><div class="label">Total Leads</div><div class="value">{total_leads}</div></div>
  <div class="stat-card"><div class="label">Qualified</div><div class="value">{qualified_count}</div></div>
  <div class="stat-card"><div class="label">Avg Score</div><div class="value">{avg_score}</div></div>
  <div class="stat-card"><div class="label">Contacts</div><div class="value">{total_contacts}</div></div>
  <div class="stat-card"><div class="label">Reports</div><div class="value">{total_reports}</div></div>
</div>
<table id="leadTable">
<thead><tr>
  <th onclick="sortTable(0)">ID</th>
  <th onclick="sortTable(1)">Company</th>
  <th onclick="sortTable(2)">Industry</th>
  <th onclick="sortTable(3)">Score</th>
  <th onclick="sortTable(4)">Confidence</th>
  <th onclick="sortTable(5)">Qualified</th>
  <th onclick="sortTable(6)">Contacts</th>
  <th>Details</th>
</tr></thead>
<tbody>
{lead_rows}
</tbody>
</table>
<script>
function sortTable(col) {{
  var table = document.getElementById("leadTable");
  var rows = Array.from(table.tBodies[0].rows).filter(r => !r.classList.contains("details"));
  var asc = table.getAttribute("data-sort-col") == col && table.getAttribute("data-sort-dir") != "asc";
  rows.sort(function(a, b) {{
    var va = a.cells[col].getAttribute("data-val") || a.cells[col].textContent;
    var vb = b.cells[col].getAttribute("data-val") || b.cells[col].textContent;
    var na = parseFloat(va), nb = parseFloat(vb);
    if (!isNaN(na) && !isNaN(nb)) return asc ? na - nb : nb - na;
    return asc ? va.localeCompare(vb) : vb.localeCompare(va);
  }});
  var tbody = table.tBodies[0];
  rows.forEach(function(r) {{
    var detailId = "detail-" + r.getAttribute("data-id");
    var detailRow = document.getElementById(detailId);
    tbody.appendChild(r);
    if (detailRow) tbody.appendChild(detailRow);
  }});
  table.setAttribute("data-sort-col", col);
  table.setAttribute("data-sort-dir", asc ? "asc" : "desc");
}}
function toggleDetail(id) {{
  var el = document.getElementById("detail-" + id);
  if (el) el.classList.toggle("open");
}}
</script>
</body>
</html>
"""


def _html_lead_row(lead: ExportLead) -> str:
    """Render a single lead as a table row + expandable detail row."""
    score_cls = "score-high" if lead.lead_score >= 0.7 else (
        "score-low" if lead.lead_score < 0.4 else ""
    )
    qual_cls = "qualified" if lead.is_qualified else "not-qualified"
    cid = lead.company_id

    main_row = (
        f'<tr data-id="{cid}">'
        f'<td data-val="{cid}">{cid}</td>'
        f'<td>{_esc(lead.company_name)}</td>'
        f'<td>{_esc(lead.industry or "")}</td>'
        f'<td data-val="{lead.lead_score:.4f}" class="{score_cls}">{lead.lead_score:.4f}</td>'
        f'<td data-val="{lead.lead_confidence:.4f}">{lead.lead_confidence:.4f}</td>'
        f'<td class="{qual_cls}">{"Yes" if lead.is_qualified else "No"}</td>'
        f'<td>{len(lead.contacts)}</td>'
        f'<td><span class="toggle" onclick="toggleDetail({cid})">expand</span></td>'
        f'</tr>\n'
    )

    # Detail row
    detail_parts: List[str] = []
    if lead.conformal:
        detail_parts.append(
            f'<b>Conformal Interval:</b> [{lead.conformal.lower_bound:.4f}, '
            f'{lead.conformal.upper_bound:.4f}] (width {lead.conformal.interval_width:.4f}, '
            f'{lead.conformal.coverage_level*100:.0f}% coverage)'
        )
    if lead.shap_explanations:
        shap_str = ", ".join(
            f"{s.feature_name}: {s.shap_value:+.3f}" for s in lead.shap_explanations[:5]
        )
        detail_parts.append(f'<b>SHAP:</b> {shap_str}')
    if lead.contacts:
        contacts_html = "<b>Contacts:</b><ul>" + "".join(
            f'<li>{_esc(c.name)} &mdash; {_esc(c.role or "")} '
            f'({_esc(c.email or "")}, {_esc(c.phone or "")})</li>'
            for c in lead.contacts
        ) + "</ul>"
        detail_parts.append(contacts_html)
    if lead.report:
        r = lead.report
        report_html = '<div class="report-section">'
        report_html += f'<h4>Report (confidence {r.confidence:.2f})</h4>'
        report_html += f'<p>{_esc(r.summary)}</p>'
        if r.key_strengths:
            report_html += '<h4>Key Strengths</h4><ul>' + "".join(
                f'<li>{_esc(s)}</li>' for s in r.key_strengths
            ) + '</ul>'
        if r.growth_indicators:
            report_html += '<h4>Growth Indicators</h4><ul>' + "".join(
                f'<li>{_esc(g)}</li>' for g in r.growth_indicators
            ) + '</ul>'
        if r.risk_factors:
            report_html += '<h4>Risk Factors</h4><ul>' + "".join(
                f'<li>{_esc(rf)}</li>' for rf in r.risk_factors
            ) + '</ul>'
        report_html += f'<h4>Recommended Approach</h4><p>{_esc(r.recommended_approach)}</p>'
        report_html += '</div>'
        detail_parts.append(report_html)

    detail_content = "<br>".join(detail_parts) if detail_parts else "<em>No additional details</em>"
    detail_row = (
        f'<tr id="detail-{cid}" class="details">'
        f'<td colspan="8">{detail_content}</td></tr>\n'
    )
    return main_row + detail_row


def _esc(text: str) -> str:
    """HTML-escape a string."""
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def export_html(leads: Sequence[ExportLead],
                dest: Union[str, Path],
                filt: Optional[ExportFilter] = None,
                pagination: Optional[PaginationConfig] = None) -> int:
    """
    Export leads as a standalone HTML report with embedded CSS,
    sortable table, and expandable per-lead detail rows.

    Returns:
        Number of leads written.
    """
    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)
    if not page:
        logger.warning("HTML export: zero leads after filtering")
        return 0

    scores = [l.lead_score for l in filtered]
    html = _HTML_TEMPLATE.format(
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"),
        total_leads=len(filtered),
        qualified_count=sum(1 for l in filtered if l.is_qualified),
        avg_score=f"{sum(scores) / len(scores):.4f}" if scores else "N/A",
        total_contacts=sum(len(l.contacts) for l in filtered),
        total_reports=sum(1 for l in filtered if l.report),
        lead_rows="".join(_html_lead_row(l) for l in page),
    )

    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(html, encoding="utf-8")
    logger.info("HTML export: %d leads written to %s", len(page), dest)
    return len(page)


# ---------------------------------------------------------------------------
# Markdown Export
# ---------------------------------------------------------------------------

def export_markdown(leads: Sequence[ExportLead],
                    dest: Union[str, Path, io.StringIO],
                    filt: Optional[ExportFilter] = None,
                    pagination: Optional[PaginationConfig] = None) -> int:
    """
    Export leads as a Markdown document suitable for GitHub / documentation.

    Returns:
        Number of leads written.
    """
    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)

    lines: List[str] = []
    lines.append("# Scrapus Lead Export\n")
    lines.append(f"**Generated:** {datetime.now().isoformat()}\n")
    lines.append(f"**Total leads:** {len(filtered)}\n")
    qual = sum(1 for l in filtered if l.is_qualified)
    lines.append(f"**Qualified:** {qual} ({qual / len(filtered) * 100:.1f}%)\n" if filtered else "")
    lines.append("")

    # Summary table
    lines.append("| # | Company | Industry | Score | Confidence | Qualified |")
    lines.append("|---|---------|----------|------:|----------:|:---------:|")
    for i, lead in enumerate(page, start=1):
        q = "Yes" if lead.is_qualified else "No"
        lines.append(
            f"| {i} | {lead.company_name} | {lead.industry or '-'} "
            f"| {lead.lead_score:.4f} | {lead.lead_confidence:.4f} | {q} |"
        )
    lines.append("")

    # Per-lead details
    for lead in page:
        lines.append(f"## {lead.company_name} (ID {lead.company_id})\n")
        lines.append(f"- **Domain:** {lead.domain or 'N/A'}")
        lines.append(f"- **Industry:** {lead.industry or 'N/A'}")
        lines.append(f"- **Size:** {lead.size_tier or 'N/A'}")
        lines.append(f"- **Location:** {lead.location or 'N/A'}")
        lines.append(f"- **Lead Score:** {lead.lead_score:.4f}")
        lines.append(f"- **Confidence:** {lead.lead_confidence:.4f}")
        lines.append(f"- **Qualified:** {'Yes' if lead.is_qualified else 'No'}")
        if lead.conformal:
            lines.append(
                f"- **Conformal Interval:** [{lead.conformal.lower_bound:.4f}, "
                f"{lead.conformal.upper_bound:.4f}] "
                f"(width {lead.conformal.interval_width:.4f})"
            )
        if lead.shap_explanations:
            lines.append("- **Top SHAP features:**")
            for s in lead.shap_explanations[:5]:
                lines.append(f"  - {s.feature_name}: {s.shap_value:+.4f} (value={s.feature_value})")
        if lead.contacts:
            lines.append(f"\n### Contacts ({len(lead.contacts)})\n")
            for c in lead.contacts:
                lines.append(f"- **{c.name}** ({c.role or 'N/A'}): {c.email or '-'} / {c.phone or '-'}")
        if lead.report:
            r = lead.report
            lines.append(f"\n### Report (confidence {r.confidence:.2f})\n")
            lines.append(f"> {r.summary}\n")
            if r.key_strengths:
                lines.append("**Key Strengths:**")
                for s in r.key_strengths:
                    lines.append(f"- {s}")
            if r.growth_indicators:
                lines.append("\n**Growth Indicators:**")
                for g in r.growth_indicators:
                    lines.append(f"- {g}")
            if r.risk_factors:
                lines.append("\n**Risk Factors:**")
                for rf in r.risk_factors:
                    lines.append(f"- {rf}")
            lines.append(f"\n**Recommended Approach:** {r.recommended_approach}")
        lines.append("\n---\n")

    content = "\n".join(lines)

    close_after = False
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        fh = open(dest, "w", encoding="utf-8")
        close_after = True
    else:
        fh = dest

    try:
        fh.write(content)
    finally:
        if close_after:
            fh.close()

    logger.info("Markdown export: %d leads written", len(page))
    return len(page)


# ---------------------------------------------------------------------------
# Template-Based Customization
# ---------------------------------------------------------------------------

@dataclass
class ExportTemplate:
    """
    Customizable export template that controls which columns / sections
    appear in an export and how they are labelled.
    """
    name: str
    description: str = ""
    # Column inclusion flags
    include_conformal: bool = True
    include_shap: bool = True
    include_contacts: bool = True
    include_entities: bool = True
    include_report: bool = True
    include_source_urls: bool = True
    # Column renaming  (internal_name -> display_name)
    column_labels: Dict[str, str] = field(default_factory=dict)
    # Custom row transform applied before serialization
    row_transform: Optional[Callable[[ExportLead], Dict[str, Any]]] = None


# Pre-built templates
TEMPLATE_MINIMAL = ExportTemplate(
    name="minimal",
    description="Company name, score, qualification status only",
    include_conformal=False,
    include_shap=False,
    include_contacts=False,
    include_entities=False,
    include_report=False,
    include_source_urls=False,
)

TEMPLATE_SALES = ExportTemplate(
    name="sales",
    description="Sales team view: company, contact, report summary",
    include_conformal=False,
    include_shap=False,
    include_entities=False,
    column_labels={
        "company_name": "Company",
        "lead_score": "Score",
        "is_qualified": "Qualified?",
        "primary_contact_name": "Contact",
        "primary_contact_email": "Email",
    },
)

TEMPLATE_DATA_SCIENCE = ExportTemplate(
    name="data_science",
    description="Full data with conformal intervals and SHAP explanations",
    include_conformal=True,
    include_shap=True,
    include_contacts=True,
    include_entities=True,
    include_report=True,
    include_source_urls=True,
)


def apply_template(lead: ExportLead, template: ExportTemplate) -> Dict[str, Any]:
    """
    Flatten a lead according to the given template.
    Respects inclusion flags and column renaming.
    """
    if template.row_transform:
        return template.row_transform(lead)

    row: Dict[str, Any] = {
        "company_id": lead.company_id,
        "company_name": lead.company_name,
        "domain": lead.domain or "",
        "industry": lead.industry or "",
        "size_tier": lead.size_tier or "",
        "location": lead.location or "",
        "lead_score": round(lead.lead_score, 4),
        "lead_confidence": round(lead.lead_confidence, 4),
        "is_qualified": lead.is_qualified,
    }

    if template.include_conformal and lead.conformal:
        row["conformal_lower"] = round(lead.conformal.lower_bound, 4)
        row["conformal_upper"] = round(lead.conformal.upper_bound, 4)
        row["conformal_width"] = round(lead.conformal.interval_width, 4)

    if template.include_shap and lead.shap_explanations:
        top = max(lead.shap_explanations, key=lambda s: abs(s.shap_value))
        row["top_shap_feature"] = top.feature_name
        row["top_shap_value"] = round(top.shap_value, 4)

    if template.include_contacts and lead.contacts:
        c = lead.contacts[0]
        row["primary_contact_name"] = c.name
        row["primary_contact_email"] = c.email or ""
        row["primary_contact_phone"] = c.phone or ""
        row["primary_contact_role"] = c.role or ""
        row["contact_count"] = len(lead.contacts)

    if template.include_report and lead.report:
        row["report_summary"] = lead.report.summary
        row["report_confidence"] = round(lead.report.confidence, 4)

    if template.include_source_urls:
        row["source_urls"] = "; ".join(lead.source_urls) if lead.source_urls else ""

    # Apply column renaming
    renamed: Dict[str, Any] = {}
    for key, val in row.items():
        label = template.column_labels.get(key, key)
        renamed[label] = val
    return renamed


def export_csv_with_template(leads: Sequence[ExportLead],
                             dest: Union[str, Path, io.StringIO],
                             template: ExportTemplate,
                             filt: Optional[ExportFilter] = None,
                             pagination: Optional[PaginationConfig] = None) -> int:
    """Export CSV using a custom template for column selection and labels."""
    filtered = apply_filters(leads, filt)
    page = paginate(filtered, pagination)
    if not page:
        return 0

    rows = [apply_template(l, template) for l in page]
    fieldnames = list(rows[0].keys())

    close_after = False
    if isinstance(dest, (str, Path)):
        dest = Path(dest)
        dest.parent.mkdir(parents=True, exist_ok=True)
        fh = open(dest, "w", newline="", encoding="utf-8")
        close_after = True
    else:
        fh = dest

    try:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    finally:
        if close_after:
            fh.close()

    logger.info("Template CSV export (%s): %d rows", template.name, len(rows))
    return len(rows)


# ---------------------------------------------------------------------------
# Unified Export Dispatcher
# ---------------------------------------------------------------------------

ExportFormat = Literal["csv", "json", "jsonl", "xlsx", "html", "markdown"]


def export_leads(leads: Sequence[ExportLead],
                 dest: Union[str, Path],
                 fmt: ExportFormat = "csv",
                 filt: Optional[ExportFilter] = None,
                 pagination: Optional[PaginationConfig] = None,
                 template: Optional[ExportTemplate] = None) -> int:
    """
    Unified entry point: export leads in the specified format.

    Args:
        leads: Sequence of hydrated ExportLead records.
        dest: Output file path.
        fmt: One of csv, json, jsonl, xlsx, html, markdown.
        filt: Optional filter to narrow lead set.
        pagination: Optional pagination config.
        template: Optional template for CSV (ignored for other formats).

    Returns:
        Number of leads / rows written.
    """
    dest = Path(dest)
    dispatchers = {
        "csv": lambda: (
            export_csv_with_template(leads, dest, template, filt, pagination)
            if template else export_csv(leads, dest, filt, pagination)
        ),
        "json": lambda: export_json(leads, dest, filt, pagination),
        "jsonl": lambda: export_jsonl(leads, dest, filt),
        "xlsx": lambda: export_xlsx(leads, dest, filt, pagination),
        "html": lambda: export_html(leads, dest, filt, pagination),
        "markdown": lambda: export_markdown(leads, dest, filt, pagination),
    }

    if fmt not in dispatchers:
        raise ValueError(f"Unsupported export format: {fmt}. Choose from {list(dispatchers.keys())}")

    count = dispatchers[fmt]()
    logger.info("Export complete: format=%s, dest=%s, count=%d", fmt, dest, count)
    return count
